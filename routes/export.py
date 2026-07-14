"""
Export Routes
Handles exporting reports to Excel and CSV formats.
"""

from flask import Blueprint, request, flash, redirect, send_file
from database.database import db
from config import config
from routes.auth import login_required
from datetime import datetime
import csv
import os
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

export_bp = Blueprint('export', __name__, url_prefix='/export')


@export_bp.route('/excel', methods=['POST'])
@login_required
def export_to_excel():
    """Export orders to Excel format"""
    try:
        export_type = request.form.get('export_type', 'all').strip()
        start_date = request.form.get('start_date', '').strip()
        end_date = request.form.get('end_date', '').strip()
        
        if export_type == 'all':
            orders = db.get_all_orders()
            filename = f"orders_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        elif export_type == 'date_range':
            if not start_date or not end_date:
                flash('Please provide start and end dates.', 'error')
                return redirect(request.referrer)
            orders = db.get_orders_by_date_range(start_date, end_date)
            filename = f"orders_{start_date}_to_{end_date}.xlsx"
        elif export_type == 'pending':
            orders = db.get_orders_by_status('Pending')
            filename = f"orders_pending_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        elif export_type == 'delivered':
            orders = db.get_orders_by_status('Delivered')
            filename = f"orders_delivered_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            orders = db.get_all_orders()
            filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        
        headers = [
            'Order ID', 'Customer Name', 'Product Name', 'Quantity', 'Price',
            'Total', 'Order Date', 'Delivery Date', 'Status', 'Payment Status',
            'Tracking Number', 'Shipping Address', 'Contact Number'
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for row_num, order in enumerate(orders, 2):
            total = order['price'] * order['quantity']
            
            row_data = [
                order['order_id'],
                order['customer_name'],
                order['product_name'],
                order['quantity'],
                f"{order['price']:.2f}",
                f"{total:.2f}",
                order['order_date'],
                order['delivery_date'] or '',
                order['status'],
                order['payment_status'],
                order['tracking_number'] or '',
                order['shipping_address'],
                order['contact_number']
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        column_widths = [15, 20, 20, 12, 12, 12, 15, 15, 15, 15, 18, 25, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        exports_path = os.path.join(config.UPLOAD_FOLDER, filename)
        with open(exports_path, 'wb') as f:
            f.write(excel_file.getvalue())
        
        flash(f'Orders exported successfully to {filename}', 'success')
        return send_file(excel_file, 
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name=filename)
    
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        flash('Error exporting orders to Excel.', 'error')
        return redirect(request.referrer)


@export_bp.route('/csv', methods=['POST'])
@login_required
def export_to_csv():
    """Export orders to CSV format"""
    try:
        export_type = request.form.get('export_type', 'all').strip()
        start_date = request.form.get('start_date', '').strip()
        end_date = request.form.get('end_date', '').strip()
        
        if export_type == 'all':
            orders = db.get_all_orders()
            filename = f"orders_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        elif export_type == 'date_range':
            if not start_date or not end_date:
                flash('Please provide start and end dates.', 'error')
                return redirect(request.referrer)
            orders = db.get_orders_by_date_range(start_date, end_date)
            filename = f"orders_{start_date}_to_{end_date}.csv"
        elif export_type == 'pending':
            orders = db.get_orders_by_status('Pending')
            filename = f"orders_pending_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        elif export_type == 'delivered':
            orders = db.get_orders_by_status('Delivered')
            filename = f"orders_delivered_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        else:
            orders = db.get_all_orders()
            filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        output = StringIO()
        writer = csv.writer(output)
        
        headers = [
            'Order ID', 'Customer Name', 'Product Name', 'Quantity', 'Price',
            'Total', 'Order Date', 'Delivery Date', 'Status', 'Payment Status',
            'Tracking Number', 'Shipping Address', 'Contact Number'
        ]
        
        writer.writerow(headers)
        
        for order in orders:
            total = order['price'] * order['quantity']
            row = [
                order['order_id'],
                order['customer_name'],
                order['product_name'],
                order['quantity'],
                f"{order['price']:.2f}",
                f"{total:.2f}",
                order['order_date'],
                order['delivery_date'] or '',
                order['status'],
                order['payment_status'],
                order['tracking_number'] or '',
                order['shipping_address'],
                order['contact_number']
            ]
            writer.writerow(row)
        
        exports_path = os.path.join(config.UPLOAD_FOLDER, filename)
        with open(exports_path, 'w', newline='') as f:
            f.write(output.getvalue())
        
        csv_bytes = BytesIO(output.getvalue().encode('utf-8'))
        
        flash(f'Orders exported successfully to {filename}', 'success')
        return send_file(csv_bytes,
                        mimetype='text/csv',
                        as_attachment=True,
                        download_name=filename)
    
    except Exception as e:
        print(f"Error exporting to CSV: {e}")
        flash('Error exporting orders to CSV.', 'error')
        return redirect(request.referrer)